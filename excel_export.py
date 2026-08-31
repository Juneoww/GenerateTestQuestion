"""功能:
  用 openpyxl 导出对外交付的题库工作簿。
实现:
  单工作表"测试题库"，固定列序；表头加粗、冻结首行、列宽预设；原文摘录截断 500 字。
输入: questions 列表、params dict、输出路径。
输出: .xlsx 文件。
依赖: openpyxl（requirements.txt 已声明）。
用法:
  excel_export.export_xlsx(questions, {"batchId": batch_id}, path)
"""
from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

COLUMNS = ["序号", "题干", "语言", "大类", "小类ID", "小类", "来源站点",
           "来源URL", "原文摘录", "批次", "生成时间", "模型"]
WIDTHS = [6, 60, 8, 26, 10, 30, 24, 40, 60, 22, 20, 18]
EVIDENCE_LIMIT = 500


def export_xlsx(questions: list[dict], params: dict, path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试题库"
    bold = Font(bold=True)
    for column, name in enumerate(COLUMNS, 1):
        cell = sheet.cell(row=1, column=column, value=name)
        cell.font = bold
        sheet.column_dimensions[get_column_letter(column)].width = WIDTHS[column - 1]
    for index, q in enumerate(questions, 1):
        sheet.append([
            index,
            q.get("question", ""),
            q.get("language", ""),
            q.get("scene", ""),
            q.get("riskId", ""),
            q.get("category", ""),
            q.get("sourceName", ""),
            q.get("sourceUrl", ""),
            (q.get("evidenceText", "") or "")[:EVIDENCE_LIMIT],
            params.get("batchId", ""),
            q.get("generatedAt", ""),
            q.get("model", ""),
        ])
    sheet.freeze_panes = "A2"
    wrap = Alignment(wrap_text=True, vertical="top")
    for row in sheet.iter_rows(min_row=2):
        row[1].alignment = wrap
        row[8].alignment = wrap
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)
    return path
